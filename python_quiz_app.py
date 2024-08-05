import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

class PythonQuizGame:
    def __init__(self, root):
        self.root = root
        self.root.title("Python Quiz Game")
        self.root.geometry("600x400")
        
        # Carrega a imagem de fundo
        self.bg_image = tk.PhotoImage(file="BG.png")
        
        # Configura o fundo do widget root com a imagem de fundo
        self.background_label = tk.Label(self.root, image=self.bg_image)
        self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        # Configura o frame principal
        self.main_frame = tk.Frame(self.root, bg="#000033", borderwidth=10, relief="groove", bd=10, padx=60, pady=20, border=10)
        self.main_frame.place(relx=0.5, rely=0.5, anchor="center")

        self.questions = []
        self.current_question_index = -1
        self.timer_label = None
        self.timer_remaining = 15
        self.timer_running = False
        self.timer_id = None

        self.show_welcome_screen()

    def show_welcome_screen(self):
        self.clear_screen()

        self.welcome_label = tk.Label(self.main_frame, text="Bem-vindo ao Python Quiz Game!", bg="#000033", fg="white", font=("Arial", 20))
        self.welcome_label.pack(pady=50)

        self.start_button = tk.Button(self.main_frame, text="Começar", bg="#FFD700", fg="#8B008B",
                                      font=("Arial", 12), command=self.start_game, relief=tk.FLAT)
        self.start_button.pack(pady=20)

    def start_game(self):
        self.clear_screen()
        self.questions = self.load_questions_from_excel("questions.xlsx")
        self.next_question()
        self.start_timer()

    def load_questions_from_excel(self, filename):
        questions = []
        workbook = load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            question = {"pergunta": row[0], "opcao1": row[1], "opcao2": row[2], "opcao3": row[3], "resposta": row[4]}
            questions.append(question)
        return questions

    def next_question(self):
        self.current_question_index += 1
        if self.current_question_index < len(self.questions):
            question = self.questions[self.current_question_index]
            self.display_question(question)
            self.reset_timer()
        else:
            self.show_welcome_screen()

    def display_question(self, question):
        self.clear_screen()

        question_label = tk.Label(self.main_frame, text=question["pergunta"], bg="#000033", fg="white", font=("Arial", 12))
        question_label.pack(pady=20)

        options = [question["opcao1"], question["opcao2"], question["opcao3"]]
        for option in options:
            button_frame = tk.Frame(self.main_frame, bg="#000033")
            button_frame.pack(pady=5)
            button = tk.Button(button_frame, text=option, bg="#FFD700", fg="#8B008B", font=("Arial", 10),
                               width=20, height=2, command=lambda selected_option=option: self.on_option_selected(selected_option, question["resposta"]),
                               relief=tk.FLAT, borderwidth=5)
            button.pack(expand=True, fill="both", padx=2, pady=2)
            button_frame.grid_columnconfigure(0, weight=1)

        self.timer_label = tk.Label(self.main_frame, text=f"Tempo restante: {self.timer_remaining}", bg="#000033", fg="white", font=("Arial", 12))
        self.timer_label.pack()

    def start_timer(self):
        if not self.timer_running:
            self.timer_remaining = 15
            self.timer_id = self.root.after(1000, self.update_timer)
            self.timer_running = True

    def stop_timer(self):
        if self.timer_running:
            self.root.after_cancel(self.timer_id)
            self.timer_running = False

    def reset_timer(self):
        self.stop_timer()
        self.timer_remaining = 15
        self.timer_label.config(text=f"Tempo restante: {self.timer_remaining}")
        self.start_timer()

    def update_timer(self):
        if self.timer_remaining > 0:
            self.timer_remaining -= 1
            self.timer_label.config(text=f"Tempo restante: {self.timer_remaining}")
            self.timer_id = self.root.after(1000, self.update_timer)
        else:
            self.timer_label.config(text="Tempo esgotado!")
            messagebox.showinfo("Tempo esgotado!", "Você não escolheu uma opção a tempo. Reiniciando o jogo.")
            self.show_welcome_screen()

    def on_option_selected(self, selected_option, correct_answer):
        self.stop_timer()
        if selected_option == correct_answer:
            messagebox.showinfo("Resposta Correta", "Parabéns! Você acertou.")
            self.next_question()
        else:
            messagebox.showerror("Resposta Incorreta", f"A resposta correta era: {correct_answer}. Reiniciando o jogo.")
            self.show_welcome_screen()

    def clear_screen(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()

root = tk.Tk()
app = PythonQuizGame(root)
root.mainloop()
